"""
Excel Report Generator
Produces a 3-sheet Excel workbook with formatting, rankings, and highlights
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel


# ─────────────────── Color Palette ───────────────────
GOLD_BG    = "FFD700"
SILVER_BG  = "C0C0C0"
BRONZE_BG  = "CD7F32"
HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
ALT_ROW    = "EBF3FF"
BEST_BG    = "00B050"
BEST_FG    = "FFFFFF"
ACCENT_BG  = "2E75B6"
ACCENT_FG  = "FFFFFF"

RANK_COLORS = {1: GOLD_BG, 2: SILVER_BG, 3: BRONZE_BG}


def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def header_border():
    side = Side(style="medium", color=HEADER_BG)
    return Border(left=side, right=side, top=side, bottom=side)

def style_header(cell, bg=HEADER_BG, fg=HEADER_FG, size=11, bold=True):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = header_border()

def style_data(cell, bold=False, center=True, bg=None):
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center"
    )
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = thin_border()


# ═══════════════════════════════════════════════
# SHEET 1: Raw Dataset
# ═══════════════════════════════════════════════

def build_sheet1(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = "Raw Dataset"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = list(df.columns)
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=header.replace("_", " ").title())
        style_header(cell)
        ws.column_dimensions[get_column_letter(c_idx)].width = max(16, len(header) + 2)

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        bg = ALT_ROW if r_idx % 2 == 0 else None
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            style_data(cell, bg=bg)

    ws.row_dimensions[1].height = 36

    # Color scale on accuracy_score column
    acc_col_idx = headers.index("accuracy_score") + 1
    acc_col = get_column_letter(acc_col_idx)
    ws.conditional_formatting.add(
        f"{acc_col}2:{acc_col}{len(df)+1}",
        ColorScaleRule(
            start_type="min", start_color="FF0000",
            mid_type="percentile", mid_value=50, mid_color="FFFF00",
            end_type="max", end_color="00B050"
        )
    )


# ═══════════════════════════════════════════════
# SHEET 2: Model Results
# ═══════════════════════════════════════════════

def build_sheet2(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Model Results")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    target_cols = [
        "accuracy_score", "adaptation_score", "task_success_rate",
        "relational_learning_score", "collaboration_efficiency"
    ]
    models = df["model_type"].unique()

    # Build summary
    summary_rows = []
    for model in ["CognitiveMorph", "Transformer", "RL", "GNN", "Symbolic"]:
        mdf = df[df["model_type"] == model]
        row = {"Model": model}
        for col in target_cols:
            row[f"{col}_mean"] = round(mdf[col].mean(), 4)
            row[f"{col}_std"]  = round(mdf[col].std(), 4)
        row["Overall_Mean"] = round(np.mean([row[f"{c}_mean"] for c in target_cols]), 4)
        summary_rows.append(row)
    sumdf = pd.DataFrame(summary_rows).sort_values("Overall_Mean", ascending=False)

    headers = (
        ["Model"] +
        [f"{c.replace('_',' ').title()} Mean" for c in target_cols] +
        [f"{c.replace('_',' ').title()} Std" for c in target_cols] +
        ["Overall Mean"]
    )
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        style_header(cell)
        ws.column_dimensions[get_column_letter(c_idx)].width = 20

    best_overall = sumdf["Overall_Mean"].max()
    for r_idx, (_, row) in enumerate(sumdf.iterrows(), 2):
        is_best = row["Overall_Mean"] == best_overall
        for c_idx, col in enumerate(headers, 1):
            key_map = {
                "Model": "Model",
                **{f"{c.replace('_',' ').title()} Mean": f"{c}_mean" for c in target_cols},
                **{f"{c.replace('_',' ').title()} Std":  f"{c}_std"  for c in target_cols},
                "Overall Mean": "Overall_Mean"
            }
            val = row.get(key_map.get(col, col), "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if is_best:
                cell.font = Font(name="Arial", size=10, bold=True, color=BEST_FG)
                cell.fill = PatternFill("solid", fgColor=BEST_BG)
                cell.border = thin_border()
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                style_data(cell, bg=ALT_ROW if r_idx % 2 == 0 else None)

    # Annotation for best model
    ws.cell(row=1, column=len(headers)+2, value="★ GREEN = Best Performing Model")
    ws.cell(row=1, column=len(headers)+2).font = Font(name="Arial", color=BEST_BG, bold=True)
    ws.row_dimensions[1].height = 40


# ═══════════════════════════════════════════════
# SHEET 3: Comparison Table
# ═══════════════════════════════════════════════

def build_sheet3(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Comparison Table")
    ws.sheet_view.showGridLines = False

    target_cols = [
        "accuracy_score", "adaptation_score", "task_success_rate",
        "relational_learning_score", "collaboration_efficiency"
    ]
    MODEL_ORDER = ["CognitiveMorph", "Transformer", "RL", "GNN", "Symbolic"]

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Multi-Architecture AI Benchmark — Comparison Table"
    title_cell.font = Font(name="Arial", size=14, bold=True, color=HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Column headers
    col_headers = ["Rank", "Model", "Accuracy", "Adaptation",
                   "Task Success", "Relational Score", "Collaboration", "Mean Score"]
    for c_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=2, column=c_idx, value=h)
        style_header(cell, bg=ACCENT_BG, fg=ACCENT_FG)
        ws.column_dimensions[get_column_letter(c_idx)].width = 18

    # Compute summary
    comparison_rows = []
    for model in MODEL_ORDER:
        mdf = df[df["model_type"] == model]
        means = {col: round(mdf[col].mean(), 4) for col in target_cols}
        means["Mean Score"] = round(np.mean(list(means.values())), 4)
        means["Model"] = model
        comparison_rows.append(means)

    comparison_df = (
        pd.DataFrame(comparison_rows)
        .sort_values("Mean Score", ascending=False)
        .reset_index(drop=True)
    )
    comparison_df.insert(0, "Rank", range(1, len(comparison_df) + 1))

    for r_idx, (_, row) in enumerate(comparison_df.iterrows(), 3):
        rank = int(row["Rank"])
        rank_bg = RANK_COLORS.get(rank)
        for c_idx, col in enumerate(col_headers, 1):
            col_key_map = {
                "Rank": "Rank", "Model": "Model",
                "Accuracy": "accuracy_score",
                "Adaptation": "adaptation_score",
                "Task Success": "task_success_rate",
                "Relational Score": "relational_learning_score",
                "Collaboration": "collaboration_efficiency",
                "Mean Score": "Mean Score"
            }
            val = row.get(col_key_map[col], "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)

            if rank == 1:
                cell.font = Font(name="Arial", bold=True, size=11, color="000000")
                cell.fill = PatternFill("solid", fgColor=GOLD_BG)
            elif rank == 2:
                cell.font = Font(name="Arial", bold=True, size=10, color="000000")
                cell.fill = PatternFill("solid", fgColor=SILVER_BG)
            elif rank == 3:
                cell.font = Font(name="Arial", size=10)
                cell.fill = PatternFill("solid", fgColor=BRONZE_BG)
            else:
                style_data(cell, bg=ALT_ROW if r_idx % 2 == 0 else None)

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()

    ws.row_dimensions[2].height = 30

    # Legend
    legend_row = len(comparison_df) + 5
    ws.cell(row=legend_row, column=1, value="LEGEND").font = Font(bold=True, name="Arial")
    ws.cell(row=legend_row+1, column=1, value="🥇 Gold = Rank 1")
    ws.cell(row=legend_row+2, column=1, value="🥈 Silver = Rank 2")
    ws.cell(row=legend_row+3, column=1, value="🥉 Bronze = Rank 3")

    # Conditional color scale on Mean Score column
    mean_col = get_column_letter(col_headers.index("Mean Score") + 1)
    ws.conditional_formatting.add(
        f"{mean_col}3:{mean_col}{len(comparison_df)+2}",
        ColorScaleRule(
            start_type="min", start_color="FF0000",
            end_type="max", end_color=BEST_BG
        )
    )


# ═══════════════════════════════════════════════
# MAIN: Build Workbook
# ═══════════════════════════════════════════════

def generate_excel_report(
    df: pd.DataFrame,
    output_path: str = "evaluation/CognitiveMorph_Benchmark_Report.xlsx"
):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()

    build_sheet1(wb, df)
    build_sheet2(wb, df)
    build_sheet3(wb, df)

    wb.save(output_path)
    print(f"\n✓ Excel report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    from data.generate_dataset import generate_dataset
    os.makedirs("data", exist_ok=True)
    os.makedirs("evaluation", exist_ok=True)
    df = generate_dataset()
    generate_excel_report(df)