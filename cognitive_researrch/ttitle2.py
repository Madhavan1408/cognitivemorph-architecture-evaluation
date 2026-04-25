"""
================================================================================
Title 2: Repercussions of Rogue Access Management Practices on Security and
         Accountability in Small and Medium Enterprises using
         Zero-Trust Access Algorithm Comparing Multi-Factor Authentication (MFA)

Problem      : Rogue and unauthorized access practices in SMEs
Intervention : Zero-Trust Access Algorithm          (Group 1, n=10)
Comparison   : Multi-Factor Authentication (MFA)    (Group 2, n=10)
Outcome      : Security Breach Reduction Rate (%)
Total Sample : 20  (2 groups × 10)
Sampling     : Stratified (≈70% safe / 30% breach) — reflects real SME exposure
================================================================================
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.ticker import MaxNLocator
from sklearn.metrics import (accuracy_score, precision_score, recall_score,
                             f1_score, confusion_matrix)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ══════════════════════════════════════════════════════════════════════════════
# PALETTE & STYLE
# ══════════════════════════════════════════════════════════════════════════════
C = {
    'zt'    : '#0D47A1',
    'mfa'   : '#B71C1C',
    'zt_lt' : '#BBDEFB',
    'mfa_lt': '#FFCDD2',
    'safe'  : '#2E7D32',
    'breach': '#C62828',
    'bg'    : '#EEF2F7',
    'panel' : '#FFFFFF',
    'grid'  : '#CFD8DC',
    'text'  : '#1A237E',
    'sub'   : '#546E7A',
    'gold'  : '#F57F17',
    'warn'  : '#E65100',
}

plt.rcParams.update({
    'font.family'    : 'DejaVu Sans',
    'axes.facecolor' : C['panel'],
    'figure.facecolor': C['bg'],
    'axes.edgecolor' : C['grid'],
    'axes.labelcolor': C['text'],
    'xtick.color'    : C['sub'],
    'ytick.color'    : C['sub'],
})

# ══════════════════════════════════════════════════════════════════════════════
# 1. LOAD DATA
# ══════════════════════════════════════════════════════════════════════════════
FILE   = 'SME_Cybersecurity_Datasets.xlsx'
study2 = pd.read_excel(FILE, sheet_name='Study2')
study3 = pd.read_excel(FILE, sheet_name='Study3')

# ══════════════════════════════════════════════════════════════════════════════
# 2. STRATIFIED SAMPLE  n=10 per group (7 safe + 3 breach/critical)
#    → mirrors realistic SME access-event distribution
# ══════════════════════════════════════════════════════════════════════════════
SEED = 42

# GROUP 1 — Zero-Trust
zt_pool        = study3[study3['algorithm'] == 'ZeroTrust'].copy()
zt_safe_pool   = zt_pool[zt_pool['security_status'] == 'safe']
zt_breach_pool = zt_pool[zt_pool['security_status'].isin(['breach','critical'])]
zt_data = pd.concat([
    zt_safe_pool.sample(7,  random_state=SEED),
    zt_breach_pool.sample(3, random_state=SEED),
]).sample(frac=1, random_state=SEED).reset_index(drop=True)
zt_data['breach']      = zt_data['security_status'].isin(['breach','critical']).astype(int)
zt_data['participant'] = [f'ZT-{i+1:02d}' for i in range(10)]

# GROUP 2 — MFA
mfa_pool        = study2[study2['auth_method'] == 'MFA'].copy()
mfa_safe_pool   = mfa_pool[mfa_pool['security_status'] == 'safe']
mfa_breach_pool = mfa_pool[mfa_pool['security_status'].isin(['breach','critical'])]
mfa_data = pd.concat([
    mfa_safe_pool.sample(7,  random_state=SEED),
    mfa_breach_pool.sample(3, random_state=SEED),
]).sample(frac=1, random_state=SEED).reset_index(drop=True)
mfa_data['breach']      = mfa_data['security_status'].isin(['breach','critical']).astype(int)
mfa_data['participant'] = [f'MFA-{i+1:02d}' for i in range(10)]

# ══════════════════════════════════════════════════════════════════════════════
# 3. OUTCOME: SECURITY BREACH REDUCTION RATE (%)
# ══════════════════════════════════════════════════════════════════════════════
def breach_stats(df):
    n        = len(df)
    breaches = int(df['breach'].sum())
    safe     = n - breaches
    rate     = (safe / n) * 100
    return n, breaches, safe, rate

zt_n,  zt_breach_cnt,  zt_safe_cnt,  zt_rate  = breach_stats(zt_data)
mfa_n, mfa_breach_cnt, mfa_safe_cnt, mfa_rate = breach_stats(mfa_data)

# ══════════════════════════════════════════════════════════════════════════════
# 4. CLASSIFICATION METRICS
#    y_true = breach label
#    y_pred = efficiency_score <= group median → predicted breach
# ══════════════════════════════════════════════════════════════════════════════
def get_metrics(df):
    y_true = df['breach'].values
    thr    = df['efficiency_score'].median()
    y_pred = (df['efficiency_score'] <= thr).astype(int)
    acc    = accuracy_score(y_true, y_pred)
    prec   = precision_score(y_true, y_pred, zero_division=0)
    rec    = recall_score(y_true, y_pred, zero_division=0)
    f1     = f1_score(y_true, y_pred, zero_division=0)
    cm     = confusion_matrix(y_true, y_pred, labels=[0,1])
    tn, fp, fn, tp = cm.ravel() if cm.size == 4 else (cm[0,0], 0, 0, 0)
    spec   = tn / (tn + fp) if (tn + fp) > 0 else 0
    npv    = tn / (tn + fn) if (tn + fn) > 0 else 0
    return dict(acc=acc, prec=prec, rec=rec, f1=f1, spec=spec, npv=npv,
                cm=cm, y_true=y_true, y_pred=y_pred, thr=thr,
                tn=tn, fp=fp, fn=fn, tp=tp)

zt_m  = get_metrics(zt_data)
mfa_m = get_metrics(mfa_data)

# ══════════════════════════════════════════════════════════════════════════════
# 5. DOMAIN METRICS
# ══════════════════════════════════════════════════════════════════════════════
zt_avg_eff   = zt_data['efficiency_score'].mean()
mfa_avg_eff  = mfa_data['efficiency_score'].mean()
zt_avg_anom  = zt_data['anomaly_score'].mean()
mfa_avg_anom = mfa_data['anomaly_score'].mean()
zt_pv        = zt_data['policy_violation_flag'].mean()  * 100
mfa_pv       = mfa_data['policy_violation_flag'].mean() * 100
zt_auth_fail  = (zt_data['authentication_status']  == 'failure').mean() * 100
mfa_auth_fail = (mfa_data['authentication_status'] == 'failure').mean() * 100

# MFA-specific
mfa_mfa_steps = mfa_data['auth_steps'].mean() if 'auth_steps' in mfa_data.columns else 0
mfa_auth_flag = mfa_data['auth_failure_flag'].mean() * 100 if 'auth_failure_flag' in mfa_data.columns else 0

# ══════════════════════════════════════════════════════════════════════════════
# 6. DETERMINE BEST ALGORITHM  (primary: BRR, secondary: avg efficiency)
# ══════════════════════════════════════════════════════════════════════════════
if zt_rate > mfa_rate:
    best_alg, best_rate, best_col = 'Zero-Trust (ZTA)', zt_rate, C['zt']
elif mfa_rate > zt_rate:
    best_alg, best_rate, best_col = 'MFA', mfa_rate, C['mfa']
else:  # tie on BRR → secondary metric: higher avg efficiency
    if zt_avg_eff >= mfa_avg_eff:
        best_alg, best_rate, best_col = 'Zero-Trust (ZTA) ★ (tie-break: efficiency)', zt_rate, C['zt']
    else:
        best_alg, best_rate, best_col = 'MFA ★ (tie-break: efficiency)', mfa_rate, C['mfa']

# ══════════════════════════════════════════════════════════════════════════════
# 7. PARTICIPANT TABLE BUILDER
# ══════════════════════════════════════════════════════════════════════════════
def make_table(df, group):
    rows = []
    for _, r in df.iterrows():
        rows.append({
            'Participant'      : r['participant'],
            'Efficiency Score' : round(float(r['efficiency_score']), 4),
            'Anomaly Score'    : round(float(r['anomaly_score']), 4),
            'Access State'     : r['access_state'],
            'Auth Status'      : r['authentication_status'],
            'Policy Violation' : int(r['policy_violation_flag']),
            'Breach (Actual)'  : int(r['breach']),
            'Security Status'  : r['security_status'],
            'Breach Reduced'   : 'Yes ✔' if r['breach'] == 0 else 'No ✘',
        })
    return pd.DataFrame(rows)

zt_tbl  = make_table(zt_data,  'ZT')
mfa_tbl = make_table(mfa_data, 'MFA')

# ══════════════════════════════════════════════════════════════════════════════
# 8. CONSOLE OUTPUT
# ══════════════════════════════════════════════════════════════════════════════
DIV = '═' * 80
print(DIV)
print('  TITLE 2 — ALGORITHM PERFORMANCE ANALYSIS')
print('  Zero-Trust Access Algorithm  vs  Multi-Factor Authentication (MFA)')
print('  Outcome: Security Breach Reduction Rate (%)')
print(DIV)

print('\n📋 GROUP 1 — Zero-Trust Access Algorithm  (n=10)\n')
print(zt_tbl.to_string(index=False))

print('\n📋 GROUP 2 — Multi-Factor Authentication (MFA)  (n=10)\n')
print(mfa_tbl.to_string(index=False))

print(f'\n{"─"*80}')
print(f'  PERFORMANCE SUMMARY  (Total n = 20)')
print(f'{"─"*80}')
print(f'  {"Metric":<42}  {"Zero-Trust (ZTA)":>16}  {"MFA":>12}')
print(f'  {"─"*42}  {"─"*16}  {"─"*12}')
metrics_print = [
    ('Security Breach Reduction Rate (%)',  f'{zt_rate:.1f}%',          f'{mfa_rate:.1f}%'),
    ('Total Participants (n)',              '10',                       '10'),
    ('Breaches',                            str(zt_breach_cnt),         str(mfa_breach_cnt)),
    ('Safe Outcomes',                       str(zt_safe_cnt),           str(mfa_safe_cnt)),
    ('Classification Accuracy (%)',         f'{zt_m["acc"]*100:.1f}%',  f'{mfa_m["acc"]*100:.1f}%'),
    ('Precision',                           f'{zt_m["prec"]:.4f}',      f'{mfa_m["prec"]:.4f}'),
    ('Recall (Sensitivity)',                f'{zt_m["rec"]:.4f}',       f'{mfa_m["rec"]:.4f}'),
    ('Specificity',                         f'{zt_m["spec"]:.4f}',      f'{mfa_m["spec"]:.4f}'),
    ('F1-Score',                            f'{zt_m["f1"]:.4f}',        f'{mfa_m["f1"]:.4f}'),
    ('Avg Efficiency Score',                f'{zt_avg_eff:.4f}',        f'{mfa_avg_eff:.4f}'),
    ('Avg Anomaly Score',                   f'{zt_avg_anom:.4f}',       f'{mfa_avg_anom:.4f}'),
    ('Policy Violation Rate (%)',           f'{zt_pv:.1f}%',            f'{mfa_pv:.1f}%'),
    ('Auth Failure Rate (%)',               f'{zt_auth_fail:.1f}%',     f'{mfa_auth_fail:.1f}%'),
]
for m, v1, v2 in metrics_print:
    print(f'  {m:<42}  {v1:>16}  {v2:>12}')

print(f'\n{DIV}')
print(f'  ★  BEST ALGORITHM  →  {best_alg}')
print(f'  ★  Breach Reduction Rate  →  {best_rate:.1f}%')
print(f'{DIV}\n')

# ══════════════════════════════════════════════════════════════════════════════
# 9.  FIGURE  —  3 × 3  DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
fig = plt.figure(figsize=(21, 17), facecolor=C['bg'])

fig.text(0.5, 0.988,
         'Repercussions of Rogue Access Management Practices on Security &'
         ' Accountability in SMEs',
         ha='center', va='top', fontsize=15, fontweight='bold', color=C['text'])
fig.text(0.5, 0.971,
         'Zero-Trust Access Algorithm (Group 1)  vs  '
         'Multi-Factor Authentication — MFA (Group 2)  |  n = 20  (10 per group)',
         ha='center', va='top', fontsize=10, color=C['sub'])

gs = gridspec.GridSpec(3, 3, figure=fig,
                       hspace=0.50, wspace=0.38,
                       left=0.06, right=0.97,
                       top=0.950, bottom=0.055)

ax_brr   = fig.add_subplot(gs[0, 0])
ax_acc   = fig.add_subplot(gs[0, 1])
ax_multi = fig.add_subplot(gs[0, 2])
ax_trend = fig.add_subplot(gs[1, :2])
ax_pie   = fig.add_subplot(gs[1, 2])
ax_cm_zt = fig.add_subplot(gs[2, 0])
ax_cm_mf = fig.add_subplot(gs[2, 1])
ax_dom   = fig.add_subplot(gs[2, 2])

def style(ax, title, ylabel='', xlabel=''):
    ax.set_facecolor(C['panel'])
    ax.set_title(title, fontsize=9.5, fontweight='bold',
                 color=C['text'], pad=9)
    if ylabel: ax.set_ylabel(ylabel, fontsize=8.5, color=C['sub'])
    if xlabel: ax.set_xlabel(xlabel, fontsize=8.5, color=C['sub'])
    ax.spines[['top','right']].set_visible(False)
    ax.spines[['left','bottom']].set_color(C['grid'])
    ax.tick_params(labelsize=8.5, colors=C['sub'])
    ax.yaxis.grid(True, color=C['grid'], linestyle='--', alpha=0.55, zorder=0)
    ax.set_axisbelow(True)

XLABELS = ['Zero-Trust\n(ZTA)', 'MFA']

# ── ① Breach Reduction Rate ──────────────────────────────────────────────────
vals = [zt_rate, mfa_rate]
grad_colors = [C['zt'], C['mfa']]
bars = ax_brr.bar(XLABELS, vals, color=grad_colors,
                  width=0.45, edgecolor='white', linewidth=2, zorder=3)
for b, v in zip(bars, vals):
    ax_brr.text(b.get_x()+b.get_width()/2, b.get_height()+1.5,
                f'{v:.1f}%', ha='center', fontsize=13,
                fontweight='bold', color=C['text'])
wi = 0 if zt_rate >= mfa_rate else 1
ax_brr.annotate('★ BEST', xy=(bars[wi].get_x()+bars[wi].get_width()/2,
                               vals[wi]+8),
                ha='center', fontsize=8.5, fontweight='bold',
                color=C['gold'],
                bbox=dict(boxstyle='round,pad=0.25', fc=C['panel'],
                          ec=C['gold'], lw=1.2))
ax_brr.set_ylim(0, 125)
style(ax_brr, '① Security Breach Reduction Rate (%)', ylabel='Rate (%)')

# ── ② Accuracy ───────────────────────────────────────────────────────────────
a_vals = [zt_m['acc']*100, mfa_m['acc']*100]
bars2  = ax_acc.bar(XLABELS, a_vals, color=[C['zt'], C['mfa']],
                    width=0.45, edgecolor='white', linewidth=2, zorder=3)
for b, v in zip(bars2, a_vals):
    ax_acc.text(b.get_x()+b.get_width()/2, b.get_height()+1.5,
                f'{v:.1f}%', ha='center', fontsize=13,
                fontweight='bold', color=C['text'])
ax_acc.set_ylim(0, 125)
style(ax_acc, '② Classification Accuracy (%)', ylabel='Accuracy (%)')

# ── ③ Multi-Metric Grouped Bar ───────────────────────────────────────────────
met_lbl  = ['Precision', 'Recall', 'Specificity', 'F1-Score']
zt_vals2 = [zt_m['prec'], zt_m['rec'], zt_m['spec'], zt_m['f1']]
mfa_vals2= [mfa_m['prec'],mfa_m['rec'],mfa_m['spec'],mfa_m['f1']]
x  = np.arange(len(met_lbl))
bw = 0.33
b1 = ax_multi.bar(x-bw/2, zt_vals2,  bw, color=C['zt'],
                  label='ZTA', edgecolor='white', linewidth=1.5, zorder=3)
b2 = ax_multi.bar(x+bw/2, mfa_vals2, bw, color=C['mfa'],
                  label='MFA', edgecolor='white', linewidth=1.5, zorder=3)
for b, v in list(zip(b1, zt_vals2)) + list(zip(b2, mfa_vals2)):
    if v > 0:
        ax_multi.text(b.get_x()+b.get_width()/2, v+0.01,
                      f'{v:.2f}', ha='center', fontsize=7, color=C['text'])
ax_multi.set_xticks(x)
ax_multi.set_xticklabels(met_lbl, fontsize=8)
ax_multi.set_ylim(0, 1.30)
ax_multi.legend(fontsize=8.5, framealpha=0.7)
style(ax_multi, '③ Classification Metrics Comparison', ylabel='Score')

# ── ④ Per-Participant Efficiency Score Trend ──────────────────────────────────
pts = list(range(1, 11))
eff_zt  = zt_data.sort_values('participant')['efficiency_score'].values
eff_mfa = mfa_data.sort_values('participant')['efficiency_score'].values
br_zt   = zt_data.sort_values('participant')['breach'].values
br_mfa  = mfa_data.sort_values('participant')['breach'].values

ax_trend.plot(pts, eff_zt,  'o-', color=C['zt'],  lw=2.4, ms=9,
              label='Zero-Trust (ZTA)', zorder=4)
ax_trend.plot(pts, eff_mfa, 's-', color=C['mfa'], lw=2.4, ms=9,
              label='MFA', zorder=4)
ax_trend.fill_between(pts, eff_zt,  alpha=0.10, color=C['zt'])
ax_trend.fill_between(pts, eff_mfa, alpha=0.10, color=C['mfa'])

for i in range(10):
    if br_zt[i]:
        ax_trend.scatter(i+1, eff_zt[i],  marker='X', s=170,
                         color=C['breach'], zorder=6, linewidths=1.5)
    if br_mfa[i]:
        ax_trend.scatter(i+1, eff_mfa[i], marker='X', s=170,
                         color=C['breach'], zorder=6, linewidths=1.5)

ax_trend.axhline(eff_zt.mean(),  ls='--', lw=1.2, color=C['zt'],  alpha=0.55)
ax_trend.axhline(eff_mfa.mean(), ls='--', lw=1.2, color=C['mfa'], alpha=0.55)
ax_trend.text(10.15, eff_zt.mean(),  f'μ={eff_zt.mean():.3f}',
              va='center', fontsize=7.5, color=C['zt'])
ax_trend.text(10.15, eff_mfa.mean(), f'μ={eff_mfa.mean():.3f}',
              va='center', fontsize=7.5, color=C['mfa'])

ax_trend.set_xticks(pts)
ax_trend.set_xticklabels([f'P{i}' for i in pts], fontsize=8.5)
ax_trend.set_ylim(0.0, 1.12)
ax_trend.legend(handles=[
    mpatches.Patch(color=C['zt'],     label='Zero-Trust (ZTA)'),
    mpatches.Patch(color=C['mfa'],    label='Multi-Factor Auth (MFA)'),
    mpatches.Patch(color=C['breach'], label='✘ Breach Event'),
], fontsize=8.5, framealpha=0.7, loc='lower right')
style(ax_trend,
      '④ Per-Participant Efficiency Score  (dashed = group mean  |  ✘ = Breach Event)',
      ylabel='Efficiency Score', xlabel='Participant  (P1–P10 per group)')

# ── ⑤ Donut Chart ────────────────────────────────────────────────────────────
total_s = zt_safe_cnt + mfa_safe_cnt
total_b = zt_breach_cnt + mfa_breach_cnt
wedges, texts, autotexts = ax_pie.pie(
    [total_s, total_b],
    labels=['Safe', 'Breach'],
    colors=[C['safe'], C['breach']],
    autopct='%1.0f%%',
    startangle=90,
    pctdistance=0.72,
    wedgeprops={'edgecolor':'white','linewidth':3,'width':0.52},
    textprops={'fontsize':9.5, 'fontweight':'bold'},
)
for at in autotexts:
    at.set_fontsize(10); at.set_fontweight('bold'); at.set_color('white')
ax_pie.text(0, 0.07, f'n=20', ha='center', fontsize=10,
            fontweight='bold', color=C['text'])
ax_pie.text(0,-0.12, 'Total', ha='center', fontsize=8.5, color=C['sub'])
ax_pie.set_title('⑤ Breach vs Safe\n(Combined n=20)',
                 fontsize=9.5, fontweight='bold', color=C['text'], pad=8)

# ── ⑥ & ⑦ Confusion Matrices ─────────────────────────────────────────────────
def draw_cm(ax, cm, title, col):
    cmap = LinearSegmentedColormap.from_list('c', ['#FAFAFA', col])
    ax.imshow(cm, cmap=cmap, vmin=0, vmax=max(cm.max(), 1), aspect='auto')
    lbls = [['TN', 'FP'], ['FN', 'TP']]
    for i in range(2):
        for j in range(2):
            v = cm[i, j]
            ax.text(j, i, f'{lbls[i][j]}\n{v}',
                    ha='center', va='center', fontsize=13, fontweight='bold',
                    color='white' if v > cm.max()*0.45 else '#333')
    ax.set_xticks([0,1]); ax.set_yticks([0,1])
    ax.set_xticklabels(['Pred Safe', 'Pred Breach'], fontsize=8.5)
    ax.set_yticklabels(['Act. Safe', 'Act. Breach'], fontsize=8.5)
    ax.set_title(title, fontsize=9.5, fontweight='bold', color=C['text'], pad=8)
    ax.spines[:].set_color(C['grid'])
    ax.tick_params(labelsize=8.5, colors=C['sub'])

draw_cm(ax_cm_zt, zt_m['cm'],  '⑥ Confusion Matrix — Zero-Trust (ZTA)', C['zt'])
draw_cm(ax_cm_mf, mfa_m['cm'], '⑦ Confusion Matrix — MFA',              C['mfa'])

# ── ⑧ Domain Metrics Horizontal Bar ─────────────────────────────────────────
dom_lbl  = ['Avg Efficiency\nScore', 'Avg Anomaly\nScore',
             'Policy Violation\nRate', 'Auth Failure\nRate']
zt_dom   = [zt_avg_eff,  zt_avg_anom,  zt_pv/100,       zt_auth_fail/100]
mfa_dom  = [mfa_avg_eff, mfa_avg_anom, mfa_pv/100, mfa_auth_fail/100]
y  = np.arange(len(dom_lbl))
bh = 0.33
ax_dom.barh(y+bh/2, zt_dom,  bh, color=C['zt'],  label='ZTA',
            edgecolor='white', linewidth=1.5, zorder=3)
ax_dom.barh(y-bh/2, mfa_dom, bh, color=C['mfa'], label='MFA',
            edgecolor='white', linewidth=1.5, zorder=3)
for i,(v1,v2) in enumerate(zip(zt_dom, mfa_dom)):
    ax_dom.text(v1+0.005, i+bh/2, f'{v1:.2f}', va='center',
                fontsize=7, color=C['zt'])
    ax_dom.text(v2+0.005, i-bh/2, f'{v2:.2f}', va='center',
                fontsize=7, color=C['mfa'])
ax_dom.set_yticks(y); ax_dom.set_yticklabels(dom_lbl, fontsize=8.5)
ax_dom.set_xlim(0, max(max(zt_dom), max(mfa_dom)) * 1.35)
ax_dom.legend(fontsize=8.5, framealpha=0.7)
ax_dom.set_facecolor(C['panel'])
ax_dom.set_title('⑧ Domain Security Metrics',
                 fontsize=9.5, fontweight='bold', color=C['text'], pad=9)
ax_dom.spines[['top','right']].set_visible(False)
ax_dom.spines[['left','bottom']].set_color(C['grid'])
ax_dom.xaxis.grid(True, color=C['grid'], linestyle='--', alpha=0.55, zorder=0)
ax_dom.set_axisbelow(True)
ax_dom.tick_params(labelsize=8.5, colors=C['sub'])

# ── Legend chips (top-right corner) ─────────────────────────────────────────
fig.legend(handles=[
    mpatches.Patch(color=C['zt'],  label='Group 1 — Zero-Trust (ZTA)'),
    mpatches.Patch(color=C['mfa'], label='Group 2 — Multi-Factor Auth (MFA)'),
], loc='upper right', bbox_to_anchor=(0.985, 0.965),
   fontsize=8.5, framealpha=0.90, edgecolor=C['grid'])

# ── Best Algorithm Ribbon ────────────────────────────────────────────────────
fig.text(0.5, 0.013,
         f'★  BEST ALGORITHM:  {best_alg}   |   '
         f'Security Breach Reduction Rate:  {best_rate:.1f}%   |   '
         f'n = 20  (Group 1: Zero-Trust n=10  ·  Group 2: MFA n=10)',
         ha='center', va='bottom', fontsize=9.5, fontweight='bold',
         color='white',
         bbox=dict(boxstyle='round,pad=0.50',
                   facecolor=best_col, alpha=0.93))

plt.savefig('cybersecurity_title2_performance.png',
            dpi=155, bbox_inches='tight', facecolor=C['bg'])
plt.close()
print('✅ Figure saved → cybersecurity_title2_performance.png')

# ══════════════════════════════════════════════════════════════════════════════
# 10. EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
wb   = Workbook()
thin = Side(style='thin', color='B0BEC5')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_style(cell, fill_hex, font_sz=10):
    cell.font      = Font(bold=True, color='FFFFFF', size=font_sz)
    cell.fill      = PatternFill('solid', fgColor=fill_hex)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    cell.border    = bdr

def data_style(cell, is_breach=False, bold=False):
    cell.fill      = PatternFill('solid', fgColor='FFEBEE' if is_breach else 'F1F8E9')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = bdr
    cell.font      = Font(size=9, bold=bold,
                          color='C62828' if is_breach else '1B5E20')

# ── Sheet 1: Participant Data ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Participant Data (n=20)'

COLS = ['Participant','Efficiency Score','Anomaly Score',
        'Access State','Auth Status','Policy Violation',
        'Breach (Actual)','Security Status','Breach Reduced?']
COL_W = [16,17,15,16,15,17,16,16,15]

def write_group(ws, start_row, df_tbl, group_title, hex_col):
    ws.merge_cells(f'A{start_row}:I{start_row}')
    tc = ws[f'A{start_row}']
    tc.value = group_title
    hdr_style(tc, hex_col, font_sz=11)
    ws.row_dimensions[start_row].height = 22
    for c, col in enumerate(COLS, 1):
        hdr_style(ws.cell(start_row+1, c, col), hex_col)
    for i, (_, r) in enumerate(df_tbl.iterrows(), start_row+2):
        brk = r['Breach (Actual)'] == 1
        vals = [r['Participant'], r['Efficiency Score'], r['Anomaly Score'],
                r['Access State'], r['Auth Status'], r['Policy Violation'],
                r['Breach (Actual)'], r['Security Status'], r['Breach Reduced']]
        for c, v in enumerate(vals, 1):
            data_style(ws.cell(i, c, v), is_breach=brk)

write_group(ws1, 1,  zt_tbl,
            'GROUP 1 — Zero-Trust Access Algorithm (ZTA)  |  n = 10', '0D47A1')
write_group(ws1, 14, mfa_tbl,
            'GROUP 2 — Multi-Factor Authentication (MFA)  |  n = 10', 'B71C1C')

for i, w in enumerate(COL_W, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: Performance Summary ─────────────────────────────────────────────
ws2 = wb.create_sheet('Performance Summary')
ws2.merge_cells('A1:C1')
hdr_style(ws2['A1'],
          '1A237E' if zt_rate >= mfa_rate else 'B71C1C', font_sz=12)
ws2['A1'].value = (
    f'PERFORMANCE SUMMARY  |  Title 2: Zero-Trust vs MFA  |  n = 20  '
    f'|  BEST: {best_alg}  ({best_rate:.1f}% BRR)'
)
ws2.row_dimensions[1].height = 26

for c, h in enumerate(['Metric','Zero-Trust (ZTA)','MFA'], 1):
    hdr_style(ws2.cell(2, c, h), '37474F')

summary_rows = [
    ('Security Breach Reduction Rate (%)', f'{zt_rate:.2f}%',          f'{mfa_rate:.2f}%'),
    ('Total Participants (n)',             '10',                       '10'),
    ('Breach Events',                      str(zt_breach_cnt),         str(mfa_breach_cnt)),
    ('Safe Outcomes',                      str(zt_safe_cnt),           str(mfa_safe_cnt)),
    ('Classification Accuracy (%)',        f'{zt_m["acc"]*100:.2f}%',  f'{mfa_m["acc"]*100:.2f}%'),
    ('Precision',                          f'{zt_m["prec"]:.4f}',      f'{mfa_m["prec"]:.4f}'),
    ('Recall (Sensitivity)',               f'{zt_m["rec"]:.4f}',       f'{mfa_m["rec"]:.4f}'),
    ('Specificity',                        f'{zt_m["spec"]:.4f}',      f'{mfa_m["spec"]:.4f}'),
    ('F1-Score',                           f'{zt_m["f1"]:.4f}',        f'{mfa_m["f1"]:.4f}'),
    ('Avg Efficiency Score',               f'{zt_avg_eff:.4f}',        f'{mfa_avg_eff:.4f}'),
    ('Avg Anomaly Score',                  f'{zt_avg_anom:.4f}',       f'{mfa_avg_anom:.4f}'),
    ('Policy Violation Rate (%)',          f'{zt_pv:.2f}%',            f'{mfa_pv:.2f}%'),
    ('Auth Failure Rate (%)',              f'{zt_auth_fail:.2f}%',     f'{mfa_auth_fail:.2f}%'),
    ('BEST ALGORITHM ★',
     '◀ WINNER' if zt_rate > mfa_rate or (zt_rate == mfa_rate and zt_avg_eff >= mfa_avg_eff) else '',
     '◀ WINNER' if mfa_rate > zt_rate or (mfa_rate == zt_rate and mfa_avg_eff > zt_avg_eff) else ''),
]

gold_fill = PatternFill('solid', fgColor='FFF9C4')
alt_fill  = PatternFill('solid', fgColor='F5F5F5')
win_fill  = PatternFill('solid', fgColor='FFF176')

for i, (m, v1, v2) in enumerate(summary_rows, 3):
    is_win = m.startswith('BEST')
    for c, val in enumerate([m, v1, v2], 1):
        cell = ws2.cell(i, c, val)
        cell.fill      = win_fill if is_win else (alt_fill if i%2 else gold_fill)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = bdr
        cell.font      = Font(bold=is_win, size=10,
                              color='B71C1C' if is_win else '212121')

ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 22

wb.save('cybersecurity_title2_results.xlsx')
print('✅ Excel saved   → cybersecurity_title2_results.xlsx')